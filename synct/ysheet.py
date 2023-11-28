""" Local target spreadsheet class and operations """

import pathlib

from copy import deepcopy

import pandas as pd
import numpy as np
from openpyxl import load_workbook

import synct.logger as log
from synct.tsheet import Tsheet

# Debug messages:
ACCESS_LOCAL_SPREADSHEET = 'access local target spreadsheet'
ADD_ROWS_LOCAL_SHEET = 'add rows in local target sheet: '
GOT_LOCAL_SPREADSHEET_ACCESS = 'got local target spreadsheet access'
LINK_UPDATE = 'link update: '
READ_LOCAL_SHEET = 'read local target sheet: '
UPDATE_LOCAL_SHEET = 'update local target sheet: '

# Error messages:
UNKNOWN_TARGET_FILE_TYPE = 'unknown target local spreadsheet file type'
MISSING_OR_INCORRECT_TARGET_FILE = 'missing or incorrect target local spreadsheet file'
READING_TARGET_FILE_FAILED = 'reading target local spreadsheet file failed'

WRONG_HEADER = 'wrong header in the sheet '
#WRONG_HEADER_OFFSET = 'header offset could be wrong in the sheet '

# Target local spreadsheet engines per file name extension
engines = {'.xlsx': 'openpyxl', '.xlsm': 'openpyxl', '.xltx': 'openpyxl', '.xltm': 'openpyxl'}

class Ysheet(Tsheet):   # pylint: disable=too-many-instance-attributes
    """ Target local spreadsheet class """
    sheet_length = {}   # It is needed for the delete_rows workaround.

    def __init__(self, config):
        """ Acccess the spreadsheet and read data """
        self.spreadsheet = config.spreadsheet
        self.access_spreadsheet()
        Tsheet.__init__(self, config)

    def access_spreadsheet(self):
        """ Access local target spreadsheet """
        log.debug(ACCESS_LOCAL_SPREADSHEET)
        try:
            self.engine = engines[pathlib.Path(self.spreadsheet).suffix]
            log.debug(self.engine)
        except KeyError:
            log.fatal_error(UNKNOWN_TARGET_FILE_TYPE)
        except TypeError:
            log.fatal_error(MISSING_OR_INCORRECT_TARGET_FILE)
        # Read the original Excel workbook into an openpyxl workbook object
        # It is needed to preserve the original formatting
        try:
            self.workbook = load_workbook(self.spreadsheet, data_only=False)
        except (OSError, ValueError) as exception:
            log.error(exception)
            log.fatal_error(READING_TARGET_FILE_FAILED)

    def get_sheet_data(self, sheet):
        """ Read sheet data to pandas data frame """
        log.debug(READ_LOCAL_SHEET + "'" + sheet + "'")
        # Get data sheet
        self.data[sheet] = pd.DataFrame(self.workbook[sheet].values)
        # Get header
        header_list = self.data[sheet].iloc[ \
                self.sheets_config[sheet].header_offset].values.flatten().tolist()
        # Remove header and rows above from the data frame
        self.data[sheet].drop(self.data[sheet].index[0:self.sheets_config[sheet].header_offset+1], \
                inplace=True)
        # Remove last empty rows from the data frame
        self.data[sheet]=self.data[sheet].iloc[:np.where(self.data[sheet].any(axis=1))[0][-1]+1]
        # Reindex the data frame
        self.data[sheet].reset_index(drop=True, inplace=True)
        # Remove last empty columns from the header
        while not header_list[-1]:
            last_column_index = len(header_list)-1
            if not self.data[sheet].iloc[:,last_column_index].isnull().all():
                break
            self.data[sheet].drop(columns=[last_column_index], inplace=True)
            header_list.pop()
        header_list = list(map(lambda x: x if x is not None else "", header_list))
        # Update the header in the data frame
        self.data[sheet].rename(columns=dict(zip(self.data[sheet].columns, header_list,)), \
                inplace=True)

        # It is needed for the delete_rows workaround:
        self.sheet_length[sheet] = len(self.data[sheet])

    def insert_rows(self, sheet, start_row, inserted_rows):     #pylint: disable=unused-argument
        """ Insert empty rows in the spreadsheet """
        log.debug(ADD_ROWS_LOCAL_SHEET + "'" + sheet + "'")

        # It is needed for the delete_rows workaround:
        self.sheet_length[sheet] = self.sheet_length[sheet] + inserted_rows

    def delete_rows(self, sheet, start_row, deleted_rows):
        """ Delete rows in the workbook """

        # Bug: delete_rows function leaves hyperlink data in the cell
        # https://foss.heptapod.net/openpyxl/openpyxl/-/issues/1318
        # This code solve the issue with the workaround

        # Workaround step 1:
        # Save copied hyperlinks on rows under the block of rows that will be deleted

        start = start_row + deleted_rows + 1
        num_rows = self.sheets_config[sheet].header_offset + self.sheet_length[sheet] - start + 1
        restore_hyperlink = {}

        for row in range(start, start + num_rows + 1):
            restore_hyperlink[row] = {}
            for cell in self.workbook[sheet][row]:
                if cell.hyperlink:
                    restore_hyperlink[row][cell.column] = deepcopy(cell.hyperlink)

        # Delete the rows
        self.workbook[sheet].delete_rows(start_row+1, deleted_rows)

        # Workaround step 2:
        # Restore hyperlinks
        start = start_row + 1

        for row in range(start, start + num_rows + 1):
            for cell in self.workbook[sheet][row]:
                if cell.column in restore_hyperlink[row+deleted_rows]:
                    cell.hyperlink = restore_hyperlink[row+deleted_rows][cell.column]
        self.sheet_length[sheet] = self.sheet_length[sheet] - deleted_rows

        self.save()

    def update_spreadsheet(self):
        """ Update the target spreadsheet data without header """
        for sheet in self.active_sheets:
            if len(self.data[sheet].index) > self.rows[sheet]:
                self.insert_rows(sheet, \
                        self.sheets_config[sheet].header_offset+self.rows[sheet]+2, \
                        len(self.data[sheet])-self.rows[sheet])     #pylint: disable=duplicate-code
            log.debug(UPDATE_LOCAL_SHEET + "'" + sheet + "'")
            for index, row in self.data[sheet].iterrows():
                for col_index, value in enumerate(row, start=1):
                    cell = self.workbook[sheet].cell( \
                            row=index+self.sheets_config[sheet].header_offset+2, column=col_index)
                    cell.value = value

    def update_column_with_links(self, sheet, column, link):
        """ Update the column in the target sheet with links """
        log.debug(LINK_UPDATE + 'sheet: ' + "'" + sheet + "'" + ' col: ' + str(column))
        col = self.data[sheet].columns.get_loc(column) + 1
        for index, _ in self.data[sheet].iterrows():
            cell = self.workbook[sheet].cell( \
                    row=index+self.sheets_config[sheet].header_offset+2, column=col)
            cell.hyperlink = link + str(cell.value)
        self.save()

    def save(self):
        """ Save the target spreadsheet """
        self.workbook.save(self.spreadsheet)
